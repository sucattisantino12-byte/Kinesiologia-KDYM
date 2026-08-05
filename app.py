from flask import Flask, jsonify, request, render_template, session, redirect, url_for
import json, os, threading
from datetime import datetime, timedelta

# Railway Volume: si existe /data usar ese directorio persistente
_BASE_DIR = '/data' if os.path.isdir('/data') else os.environ.get('DATA_DIR', os.path.dirname(os.path.abspath(__file__)))
os.makedirs(_BASE_DIR, exist_ok=True)
CONFIG_FILE   = os.path.join(_BASE_DIR, 'config.json')
DATA_FILE     = os.path.join(_BASE_DIR, 'data.json')
HISTORIAL_FILE = os.path.join(_BASE_DIR, 'historial.json')
VIDEOS_DIR    = os.path.join(_BASE_DIR, 'static_videos')
LOGOS_DIR     = os.path.join(_BASE_DIR, 'static_logos')
STATE_FILE    = os.path.join(_BASE_DIR, 'state.json')
os.makedirs(VIDEOS_DIR, exist_ok=True)
os.makedirs(LOGOS_DIR, exist_ok=True)

CAJA_NOMBRES = {1: 'Abajo', 2: 'Extendido', 3: 'VIP'}

DEFAULT_CONFIG = {
    'password': '1212',
    'pin_manager': '4321',
    'pin_cajaabajo': '1234',
    'pin_cajaextendido': '1234',
    'pin_cajavip': '1234',
    'pin_tarjetas': '1234',
    # Clave del panel de dueño /suca (branding + pagos). Cambiala desde el propio panel.
    'suca_password': 'suca2026',
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
                for k, v in DEFAULT_CONFIG.items():
                    if k not in cfg:
                        cfg[k] = v
                _ensure_estaciones(cfg)
                return cfg
        except Exception:
            pass
    cfg = dict(DEFAULT_CONFIG)
    _ensure_estaciones(cfg)
    return cfg

def save_config(cfg):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# ── Estaciones dinámicas (cajas / tarjetas) ──────────────────────────────────
# Cada estación: {id, tipo: 'caja'|'tarjetas', nombre, pin, activa}. El rol de
# sesión es 'est<id>'. Las 3 cajas + tarjetas originales se migran a ids 1..4
# para no romper transacciones ni historial existentes.

def _icono_default(tipo):
    return 'tarjetas' if tipo == 'tarjetas' else 'caja'

def _ensure_estaciones(cfg):
    if isinstance(cfg.get('estaciones'), list):
        # Completar 'icono' en configs viejas que no lo tengan
        for e in cfg['estaciones']:
            if not e.get('icono'):
                e['icono'] = _icono_default(e.get('tipo'))
        return cfg
    cfg['estaciones'] = [
        {'id': 1, 'tipo': 'caja',     'nombre': CAJA_NOMBRES[1], 'pin': str(cfg.get('pin_cajaabajo', '1234')),     'activa': True, 'icono': 'caja'},
        {'id': 2, 'tipo': 'caja',     'nombre': CAJA_NOMBRES[2], 'pin': str(cfg.get('pin_cajaextendido', '1234')), 'activa': True, 'icono': 'caja'},
        {'id': 3, 'tipo': 'caja',     'nombre': CAJA_NOMBRES[3], 'pin': str(cfg.get('pin_cajavip', '1234')),       'activa': True, 'icono': 'caja'},
        {'id': 4, 'tipo': 'tarjetas', 'nombre': 'Tarjetas',      'pin': str(cfg.get('pin_tarjetas', '1234')),      'activa': True, 'icono': 'tarjetas'},
    ]
    cfg['estacion_id_counter'] = 4
    return cfg

def _estaciones(cfg=None):
    return (cfg or load_config()).get('estaciones', [])

def _estacion(eid, cfg=None):
    try:
        eid = int(eid)
    except (ValueError, TypeError):
        return None
    for e in _estaciones(cfg):
        if int(e.get('id', 0)) == eid:
            return e
    return None

def _caja_nombre_map(cfg=None):
    """{id: nombre} de las estaciones tipo caja, con fallback a los nombres viejos."""
    m = dict(CAJA_NOMBRES)
    for e in _estaciones(cfg):
        if e.get('tipo') == 'caja':
            m[int(e['id'])] = e.get('nombre') or ('Caja ' + str(e['id']))
    return m

def _caja_nombre_hist(noche, cid):
    """Nombre de caja para una noche del historial: usa el snapshot guardado esa
    noche (cajas_nombres) y cae a los nombres actuales / genéricos."""
    try:
        cid = int(cid)
    except (ValueError, TypeError):
        return str(cid)
    snap = (noche or {}).get('cajas_nombres', {}) or {}
    return snap.get(str(cid)) or _caja_nombre_map().get(cid) or ('Caja ' + str(cid))

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'jagger_vip_secret_k9x_2024')
app.permanent_session_lifetime = timedelta(hours=12)
# Que los cambios en templates/ y static/ se vean sin reiniciar el programa
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
lock = threading.Lock()

_db = {'transactions': [], 'tx_id_counter': 0, 'tarjetas': {}, 'tarjetas_conf': [], 'menu': [], 'menu_id_counter': 0}

# Cargar datos guardados al iniciar
if os.path.exists(DATA_FILE):
    try:
        with open(DATA_FILE, 'r', encoding='utf-8') as _f:
            _loaded = json.load(_f)
            _db.update(_loaded)
    except Exception as _e:
        print(f'[data] Error cargando data.json: {_e}')

# Estado compartido entre todos los dispositivos
_state = {
    'hora_fin': '05:30',
    'premio': '',
    'menu_ver': 0,          # se incrementa ante cualquier cambio de menú/categoría/cartel
    'winner_show': False,
    'winner_ts': 0,
    'cartel_show': False,
    'cartel_ts': 0,
    'cartel_data': {},
    'cartel_precios': {'virtual': 0, 'fisico': 0, 'combo': 0},
    'publicidad_activa': False,
    'publicidad_url': '',
    'publicidad_frecuencia': 15,
    'publicidad_mostrar_ts': 0,
    'design': {
        'tema': 'default',
        'colores': {
            '--gold': '#c9a227', '--gold-light': '#e8c84a', '--gold-dim': '#7a6010',
            '--black': '#080808', '--surface': '#111111', '--surface-gold': '#0d0b00',
            '--border': '#2a2a2a', '--text': '#f0ece0', '--text-dim': '#555555',
            '--white': '#ffffff', '--green': '#2ecc71', '--danger': '#a83030',
        },
        'logo': 'RANKING',
        'vip': 'VIP',
        'tagline': 'JAGGER CLUB',
        'tagline_color': '#555555',
        'tagline_glow': '0',
        'tagline_font': "'Rajdhani',sans-serif",
        'winner_msg': '¡EL GANADOR DE LA NOCHE!',
        'winner_sub': '',
        'premio': '',
        'premio_size': '22',
        'hora_fin': '05:30',
        'efecto': 'ninguno',
        'tipo_particula': 'confetti',
        'deco_activa': True,
        'petals_activos': True,
        'falling_gloves': True,
        'arg_estrellas': 'doradas',
        'arg_fondo': 'claro',
        'arg_estrellas_on': True,
        # ── Sistema de diseño JAGGER 2.0 ──
        'font': 'Space Grotesk',
        'accent': 'Lima',
        'acColor': '#9ef01a',
        'efectos': 'Brillos',
        'fondo': 'Acento',
        'fondoIntensidad': 'Visible',
        'flot': 'Off',
        'modo_caja': 'dark',
        'premio_on': True,
        'relojSize': 58,
        'fx': {'confetti': True, 'grilla': False, 'animPos': True, 'sheen': True},
        'textos': {},
        'extras': [],
    },
    # ── Branding del boliche (lo edita el dueño desde /suca) ──
    'branding': {
        'color': '',       # color general de la app (acento --ac); '' = usar el del tema
        'logo_url': '',     # imagen del logo del boliche (hub, manager, pantalla)
        'nombre': '',       # nombre del boliche (opcional)
    },
    # ── Sistema de pagos del boliche al dueño (lo edita /suca; lo ve el manager/hub) ──
    'pagos': {
        'activo': False,    # si el sistema de avisos está prendido
        'vence': '',        # fecha del próximo pago 'YYYY-MM-DD'
        'monto': 0,         # monto del pago (opcional)
        'nota': '',         # nota que ve el boliche
        'dias_aviso': 5,    # días antes del vencimiento en que arranca el nivel 1
        'override': 'auto', # 'auto' (por fecha) | 'off' | '1' | '2' (forzado a mano)
    },
}

# Claves del state que se persisten en disco (excluye timestamps efímeros)
_STATE_PERSIST_KEYS = {
    'hora_fin', 'premio', 'cartel_precios', 'menu_ver',
    'publicidad_activa', 'publicidad_url', 'publicidad_frecuencia',
    'design', 'branding', 'pagos',
}

def save_state():
    """Persiste las claves importantes del _state en disco."""
    try:
        data = {k: _state[k] for k in _STATE_PERSIST_KEYS if k in _state}
        with open(STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def load_state():
    """Carga el state persistido del disco al arrancar."""
    if not os.path.exists(STATE_FILE):
        return
    try:
        with open(STATE_FILE, 'r', encoding='utf-8') as f:
            saved = json.load(f)
        for k, v in saved.items():
            if k in _state:
                # Los dicts de config se fusionan para no perder sub-claves nuevas
                if k in ('design', 'branding', 'pagos') and isinstance(_state.get(k), dict) and isinstance(v, dict):
                    _state[k].update(v)
                else:
                    _state[k] = v
    except Exception:
        pass

load_state()  # Cargar al iniciar

def load_data():
    return _db

def save_data(data):
    try:
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception as e:
        print(f'[data] Error guardando: {e}')

def bump_menu():
    """Marca que el menú/categorías/cartel cambió, para que las cajas lo detecten."""
    import time as _t
    _state['menu_ver'] = _t.time()
    save_state()

def load_historial():
    if os.path.exists(HISTORIAL_FILE):
        try:
            with open(HISTORIAL_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return []

def save_historial(h):
    try:
        with open(HISTORIAL_FILE, 'w', encoding='utf-8') as f:
            json.dump(h, f, ensure_ascii=False)
    except Exception as e:
        print(f'[historial] Error guardando: {e}')


# ══════════════════════════════════════════════════════════════════════════════
#  AUTENTICACIÓN
# ══════════════════════════════════════════════════════════════════════════════

def check_auth(role):
    if role != 'manager' and session.get('manager_auth') == True:
        cfg_v = load_config()
        if session.get('session_version', -1) == cfg_v.get('session_version', 0):
            return True
    if session.get(role + '_auth') != True:
        return False
    cfg_v = load_config()
    if session.get('session_version', -1) != cfg_v.get('session_version', 0):
        session.clear()
        return False
    return True

def require_auth(role, title, redirect_to):
    if check_auth(role):
        return None
    return render_template('login.html', title=title, role=role, redirect_to=redirect_to)


# ══════════════════════════════════════════════════════════════════════════════
#  PÁGINAS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/')
def hub_page():
    return render_template('hub.html')

@app.route('/pantalla')
def pantalla_page():
    return render_template('pantalla.html')

@app.route('/ganador')
def ganador_page():
    return render_template('ganador.html')

@app.route('/historial')
def historial_page():
    return render_template('historial.html')

@app.route('/celular')
def celular_page():
    return render_template('monitor.html')

def _serve_estacion(eid):
    e = _estacion(eid)
    if not e or not e.get('activa', True):
        return redirect(url_for('hub_page'))
    role = 'est' + str(e['id'])
    login = require_auth(role, e.get('nombre', ''), '/estacion/' + str(e['id']))
    if login: return login
    if e.get('tipo') == 'tarjetas':
        return render_template('tarjetas.html')
    return render_template('caja.html', caja_num=e['id'], caja_nombre=e.get('nombre', ''))

@app.route('/estacion/<int:eid>')
def estacion_page(eid):
    return _serve_estacion(eid)

@app.route('/tarjetas')
def tarjetas_page():
    return _serve_estacion(4)

@app.route('/manager')
def manager_page():
    login = require_auth('manager', 'Manager', '/manager')
    if login: return login
    ensure_menu()
    return render_template('manager.html')

@app.route('/pantalla-mgr')
def pantalla_mgr_page():
    login = require_auth('manager', 'Manager', '/pantalla-mgr')
    if login: return login
    return render_template('pantalla_mgr.html')

# Alias de compatibilidad con las URLs viejas (redirigen al modelo de estaciones)
@app.route('/cajaabajo')
def cajaabajo_page():
    return _serve_estacion(1)

@app.route('/cajaextendido')
def cajaextendido_page():
    return _serve_estacion(2)

@app.route('/cajavip')
def cajavip_page():
    return _serve_estacion(3)

@app.route('/api/tx', methods=['GET'])
def get_tx():
    with lock:
        data = load_data()
    return jsonify(data['transactions'])

@app.route('/api/names', methods=['GET'])
def get_names():
    data = load_data()
    names = sorted(list({t['name'] for t in data.get('transactions', []) if t.get('name')}), key=str.lower)
    return jsonify(names)

@app.route('/api/tx', methods=['POST'])
def add_tx():
    try:
        with lock:
            data = load_data()
            body = request.get_json()
            if not body or 'name' not in body or 'amount' not in body or 'caja' not in body:
                return jsonify({'ok': False, 'error': 'Faltan campos requeridos'}), 400
            amount = float(body['amount'])
            codigo = str(body.get('tarjeta_codigo',''))

            # Normalizar nombre: si ya existe alguien con ese nombre (distinto case), usar el canónico
            raw_name = str(body['name']).strip()
            existing_names = {t['name'] for t in data.get('transactions', [])}
            canonical_name = next((n for n in existing_names if n.lower() == raw_name.lower()), raw_name)

            if codigo:
                if 'tarjetas' not in data: data['tarjetas'] = {}
                if codigo not in data['tarjetas']:
                    conf_list = data.get('tarjetas_conf', [])
                    conf = next((t for t in conf_list if t.get('codigo')==codigo), None)
                    try:
                        saldo_ini = float(conf['saldo_inicial']) if conf and conf.get('saldo_inicial') not in ('', None) else 0
                    except (ValueError, TypeError):
                        saldo_ini = 0
                    data['tarjetas'][codigo] = {'saldo_actual': saldo_ini, 'nombre': canonical_name}
                saldo_disponible = data['tarjetas'][codigo]['saldo_actual']
                if saldo_disponible < amount:
                    return jsonify({'ok': False, 'error': 'Saldo insuficiente. Disponible: $' + str(int(saldo_disponible))}), 400
            data['tx_id_counter'] += 1
            tx = {
                'id': data['tx_id_counter'],
                'name': canonical_name,
                'amount': amount,
                'caja': int(body['caja']),
                'mesa': str(body.get('mesa','')),
                'tarjeta_codigo': codigo,
                'time': str(body.get('client_time','')) or datetime.now().strftime('%H:%M'),
                'items': body.get('items', []),
            }
            data['transactions'].append(tx)
            if codigo:
                data['tarjetas'][codigo]['saldo_actual'] -= amount
                data['tarjetas'][codigo]['nombre'] = canonical_name
            save_data(data)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/tx/<int:tid>', methods=['PUT'])
def edit_tx(tid):
    try:
        with lock:
            data = load_data()
            tx = next((t for t in data['transactions'] if t['id']==tid), None)
            if not tx:
                return jsonify({'ok': False, 'error': 'Transacción no encontrada'}), 404
            body = request.get_json() or {}
            new_amount = float(body.get('amount', tx['amount']))
            diff = new_amount - tx['amount']
            codigo = tx.get('tarjeta_codigo', '')
            if codigo and diff != 0 and 'tarjetas' in data and codigo in data['tarjetas']:
                nuevo_saldo = data['tarjetas'][codigo]['saldo_actual'] - diff
                if nuevo_saldo < 0:
                    return jsonify({'ok': False, 'error': 'Saldo insuficiente para este ajuste'}), 400
                data['tarjetas'][codigo]['saldo_actual'] = nuevo_saldo
            tx['amount'] = new_amount
            if 'name' in body and str(body['name']).strip():
                tx['name'] = str(body['name']).strip()
            save_data(data)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/tx/<int:tid>', methods=['DELETE'])
def delete_tx(tid):
    with lock:
        data = load_data()
        tx = next((t for t in data['transactions'] if t['id']==tid), None)
        if tx:
            codigo = tx.get('tarjeta_codigo','')
            if codigo and 'tarjetas' in data and codigo in data['tarjetas']:
                data['tarjetas'][codigo]['saldo_actual'] += tx['amount']
            data['transactions'] = [t for t in data['transactions'] if t['id']!=tid]
            save_data(data)
    return jsonify({'ok': True})

@app.route('/api/tx/<int:tid>/cancelar', methods=['POST'])
def cancelar_tx(tid):
    """Anula una operación devolviendo el saldo a la tarjeta, pero la deja
    registrada en el historial como ANULADA (no se borra)."""
    with lock:
        data = load_data()
        tx = next((t for t in data['transactions'] if t['id'] == tid), None)
        if not tx:
            return jsonify({'ok': False, 'error': 'Operación no encontrada'}), 404
        if tx.get('cancelado'):
            return jsonify({'ok': False, 'error': 'La operación ya estaba anulada'}), 400
        codigo = tx.get('tarjeta_codigo', '')
        if codigo and codigo in data.get('tarjetas', {}):
            data['tarjetas'][codigo]['saldo_actual'] += tx['amount']
        tx['cancelado'] = True
        save_data(data)
    return jsonify({'ok': True})


@app.route('/api/tarjetas', methods=['GET'])
def get_tarjetas():
    with lock:
        data = load_data()
    return jsonify(data.get('tarjetas', {}))

@app.route('/api/tarjetas/config', methods=['GET'])
def get_tarjetas_conf():
    with lock:
        data = load_data()
    return jsonify(data.get('tarjetas_conf', []))

@app.route('/api/tarjetas/config', methods=['POST'])
def set_tarjetas_conf():
    try:
        with lock:
            data = load_data()
            conf = request.get_json()
            if not conf:
                return jsonify({'ok': False, 'error': 'Configuracion invalida'}), 400
            codigos_vistos = {}
            for t in conf:
                codigo = t.get('codigo','')
                if codigo:
                    if codigo in codigos_vistos:
                        slot_anterior = codigos_vistos[codigo]
                        return jsonify({'ok': False, 'error': f'La tarjeta ya esta asignada a Mesa {slot_anterior}'}), 400
                    codigos_vistos[codigo] = t.get('slot', '?')
            conf_vieja = {t.get('codigo',''): t for t in data.get('tarjetas_conf', []) if t.get('codigo','')}
            data['tarjetas_conf'] = conf
            if 'tarjetas' not in data: data['tarjetas'] = {}
            for t in conf:
                codigo = t.get('codigo','')
                if codigo and t.get('saldo_inicial'):
                    nuevo_saldo_ini = float(t['saldo_inicial'])
                    if codigo not in data['tarjetas']:
                        data['tarjetas'][codigo] = {'saldo_actual': nuevo_saldo_ini, 'nombre':''}
                    else:
                        old_conf = conf_vieja.get(codigo)
                        old_ini = float(old_conf['saldo_inicial']) if old_conf and old_conf.get('saldo_inicial') else None
                        if old_ini is not None and old_ini != nuevo_saldo_ini:
                            gastado = old_ini - data['tarjetas'][codigo]['saldo_actual']
                            nuevo_saldo_act = max(0, nuevo_saldo_ini - gastado)
                            data['tarjetas'][codigo]['saldo_actual'] = nuevo_saldo_act
                        elif old_ini is None:
                            data['tarjetas'][codigo]['saldo_actual'] = nuevo_saldo_ini
            save_data(data)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/tarjetas/limpiar', methods=['POST'])
def limpiar_tarjetas():
    with lock:
        data = load_data()
        data['tarjetas_conf'] = []
        data['tarjetas'] = {}
        save_data(data)
    return jsonify({'ok': True})

@app.route('/api/reset', methods=['POST'])
def reset():
    """Reinicia la noche. Con {"tarjetas": "todo"} además desasigna las tarjetas
    (borra código, nombre y monto de cada mesa) para arrancar de cero."""
    body = request.get_json(silent=True) or {}
    limpiar_todo = str(body.get('tarjetas', '')) == 'todo'
    with lock:
        data = load_data()
        data['transactions'] = []
        data['tx_id_counter'] = 0
        conf_list = data.get('tarjetas_conf', [])
        data['tarjetas'] = {}
        if limpiar_todo:
            # La mesa queda vacía: hay que volver a pasar la tarjeta
            for t in conf_list:
                t['codigo'] = ''
                t['nombre'] = ''
                t['saldo_inicial'] = 0
        else:
            for t in conf_list:
                codigo = t.get('codigo', '')
                if codigo and t.get('saldo_inicial'):
                    data['tarjetas'][codigo] = {'saldo_actual': float(t['saldo_inicial']), 'nombre': ''}
        save_data(data)
    return jsonify({'ok': True})

@app.route('/api/tarjetas/nombres', methods=['POST'])
def limpiar_nombres():
    """Borra solo los nombres de los clientes; deja tarjetas, montos y consumos."""
    with lock:
        data = load_data()
        for t in data.get('tarjetas_conf', []):
            t['nombre'] = ''
            t['nombre_cliente'] = ''
        for v in data.get('tarjetas', {}).values():
            v['nombre'] = ''
        save_data(data)
    return jsonify({'ok': True})

@app.route('/api/state', methods=['GET'])
def get_state():
    return jsonify(_state)

@app.route('/api/state', methods=['POST'])
def set_state():
    import time
    body = request.get_json() or {}
    with lock:
        if 'hora_fin' in body: _state['hora_fin'] = str(body['hora_fin'])
        if 'premio' in body: _state['premio'] = str(body['premio'])
        save_state()
    return jsonify({'ok': True})

@app.route('/api/winner/show', methods=['POST'])
def winner_show():
    import time
    with lock:
        _state['winner_show'] = True
        _state['winner_ts'] = int(time.time() * 1000)
    return jsonify({'ok': True})

@app.route('/api/winner/hide', methods=['POST'])
def winner_hide():
    with lock:
        _state['winner_show'] = False
    return jsonify({'ok': True})

@app.route('/api/cartel/show', methods=['POST'])
def cartel_show():
    import time
    body = request.get_json() or {}
    with lock:
        _state['cartel_show'] = True
        _state['cartel_ts'] = int(time.time() * 1000)
        _state['cartel_data'] = {
            'nombre': str(body.get('nombre','')),
            'mesa': str(body.get('mesa','')),
            'frase': str(body.get('frase','')),
            'emoji': str(body.get('emoji','🍾')),
        }
    return jsonify({'ok': True})

@app.route('/api/cartel/hide', methods=['POST'])
def cartel_hide():
    with lock:
        _state['cartel_show'] = False
    return jsonify({'ok': True})

@app.route('/api/cartel/precios', methods=['GET'])
def get_cartel_precios():
    return jsonify(_state.get('cartel_precios', {'virtual': 0, 'fisico': 0, 'combo': 0}))

@app.route('/api/cartel/precios', methods=['POST'])
def set_cartel_precios():
    body = request.get_json() or {}
    with lock:
        if 'cartel_precios' not in _state:
            _state['cartel_precios'] = {'virtual': 0, 'fisico': 0, 'combo': 0}
        for k in ('virtual', 'fisico', 'combo'):
            if k in body:
                try:
                    _state['cartel_precios'][k] = float(body[k])
                except (ValueError, TypeError):
                    pass
        save_state()
        bump_menu()
    return jsonify({'ok': True})

@app.route('/api/cerrar_noche', methods=['POST'])
def cerrar_noche():
    with lock:
        data = load_data()
        txs = [t for t in data['transactions'] if not t.get('cancelado')]
        if not txs:
            return jsonify({'ok': False, 'error': 'No hay operaciones esta noche'}), 400
        totals, mesas, por_caja = {}, {}, {}
        for t in txs:
            totals[t['name']] = totals.get(t['name'], 0) + t['amount']
            if t.get('mesa') and t['name'] not in mesas:
                mesas[t['name']] = t['mesa']
            c = int(t.get('caja', 1))
            por_caja[c] = por_caja.get(c, 0) + t['amount']
        ranking = sorted(
            [{'name': n, 'total': v, 'mesa': mesas.get(n, '')} for n, v in totals.items()],
            key=lambda x: -x['total']
        )
        _nombres = _caja_nombre_map()
        noche = {
            'id': datetime.now().strftime('%Y%m%d_%H%M%S'),
            'fecha': datetime.now().strftime('%Y-%m-%d'),
            'hora_cierre': datetime.now().strftime('%H:%M'),
            'hora_fin': _state.get('hora_fin', '05:30'),
            'total': sum(t['amount'] for t in txs),
            'operaciones': len(txs),
            'por_caja': por_caja,
            # Snapshot de nombres de caja de esa noche (para el historial/Excel aunque después cambien)
            'cajas_nombres': {str(cid): _nombres.get(cid, 'Caja ' + str(cid)) for cid in por_caja},
            'ranking': ranking,
            'transactions': list(txs),
        }
        historial = load_historial()
        historial.append(noche)
        save_historial(historial)
    return jsonify({'ok': True, 'noche_id': noche['id']})

@app.route('/api/historial', methods=['GET'])
def get_historial():
    return jsonify(load_historial())

@app.route('/api/export/excel')
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils import get_column_letter
        from io import BytesIO
    except ImportError:
        return jsonify({'error': 'Instalá openpyxl: pip install openpyxl'}), 500

    historial = load_historial()
    if not historial:
        return jsonify({'error': 'No hay noches registradas aún'}), 400

    wb = Workbook()

    # ── Paleta única coherente ────────────────────────────────
    GOLD    = 'C9A227'
    GOLD_LT = 'F5E6A0'
    BLACK   = '111111'
    DARK    = '1E1E1E'
    WHITE   = 'FFFFFF'
    LGRAY   = 'F2F2F2'
    MGRAY   = 'D9D9D9'
    DGRAY   = '666666'
    GREEN   = '1E6B3C'
    GREEN_LT= 'D6EFDF'
    RED     = 'A83030'
    RED_LT  = 'F8D7DA'
    BLUE    = '1A4A8A'
    BLUE_LT = 'D0E4F7'

    def fill(c): return PatternFill('solid', fgColor=c)
    def side(c='CCCCCC', s='thin'): return Side(style=s, color=c)
    def border(c='CCCCCC'):
        s = side(c); return Border(left=s, right=s, top=s, bottom=s)
    def font(bold=False, color=BLACK, size=10, name='Calibri'):
        return Font(bold=bold, color=color, size=size, name=name)
    def align(h='left', v='center', wrap=False, indent=0):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

    BDR      = border('CCCCCC')
    BDR_GOLD = border(GOLD)
    BDR_DK   = border('999999')

    def cell(ws, r, c, val, f=None, bg=None, bdr=None, al=None, fmt=None):
        cc = ws.cell(row=r, column=c, value=val)
        if f:   cc.font      = f
        if bg:  cc.fill      = fill(bg)
        if bdr: cc.border    = bdr
        if al:  cc.alignment = al
        if fmt: cc.number_format = fmt
        return cc

    def row_h(ws, r, h): ws.row_dimensions[r].height = h
    def col_w(ws, c, w): ws.column_dimensions[get_column_letter(c)].width = w

    def title_row(ws, r, text, sub=''):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        cc = ws.cell(row=r, column=1, value=text)
        cc.font = Font(bold=True, color=GOLD, size=18, name='Calibri')
        cc.fill = fill(BLACK); cc.alignment = align('left', indent=2)
        row_h(ws, r, 40)
        if sub:
            ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=12)
            sc = ws.cell(row=r+1, column=1, value=sub)
            sc.font = Font(color=DGRAY, size=9, name='Calibri')
            sc.fill = fill(BLACK); sc.alignment = align('left', indent=2)
            row_h(ws, r+1, 15)

    def section(ws, r, text, col_start=1, col_end=8):
        ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_end)
        cc = ws.cell(row=r, column=col_start, value=text)
        cc.font = Font(bold=True, color=WHITE, size=10, name='Calibri')
        cc.fill = fill(DARK); cc.alignment = align('left', indent=1)
        row_h(ws, r, 20)

    def header_row(ws, r, labels, col_start=1, bg=GOLD):
        for ci, lbl in enumerate(labels, col_start):
            cc = ws.cell(row=r, column=ci, value=lbl)
            cc.font = Font(bold=True, color=BLACK, size=9, name='Calibri')
            cc.fill = fill(bg); cc.border = BDR_GOLD
            cc.alignment = align('center')
        row_h(ws, r, 20)

    def data_cell(ws, r, c, val, bold=False, alt=False, fmt=None, color=None, h='center'):
        bg = LGRAY if alt else WHITE
        f = font(bold=bold, color=color or BLACK)
        al = align(h)
        cc = ws.cell(row=r, column=c, value=val)
        cc.font = f; cc.fill = fill(bg); cc.border = BDR; cc.alignment = al
        if fmt: cc.number_format = fmt
        row_h(ws, r, 17)
        return cc

    def money(ws, r, c, val, alt=False, bold=False, color=None):
        data_cell(ws, r, c, val, bold=bold, alt=alt, fmt='#.##0', h='right', color=color or GREEN)

    def pct(ws, r, c, val, alt=False):
        data_cell(ws, r, c, val, alt=alt, fmt='0.0%', h='center')

    def kpi_block(ws, r, c, label, value, fmt='#.##0', val_color=GREEN):
        lc = ws.cell(row=r, column=c, value=label)
        lc.font = Font(bold=False, color=DGRAY, size=8, name='Calibri')
        lc.fill = fill(LGRAY); lc.border = border('E0E0E0')
        lc.alignment = align('left', indent=1); row_h(ws, r, 22)
        vc = ws.cell(row=r, column=c+1, value=value)
        vc.font = Font(bold=True, color=val_color, size=12, name='Calibri')
        vc.fill = fill(WHITE); vc.border = border('E0E0E0')
        vc.alignment = align('right', indent=1)
        if fmt: vc.number_format = fmt

    # ── Pre-cálculos ─────────────────────────────────────────
    total_all  = sum(n['total'] for n in historial)
    total_ops  = sum(n['operaciones'] for n in historial)
    avg_noche  = total_all / len(historial) if historial else 0
    mejor      = max(n['total'] for n in historial) if historial else 0

    rk_all = {}
    noches_cliente = {}
    for n in historial:
        vistos = set()
        for rv in n.get('ranking', []):
            nm = rv['name']
            rk_all[nm] = rk_all.get(nm, 0) + rv['total']
            if nm not in vistos:
                noches_cliente[nm] = noches_cliente.get(nm, 0) + 1
                vistos.add(nm)
    rk_sorted = sorted(rk_all.items(), key=lambda x: -x[1])

    por_mes = {}
    noches_mes = {}
    for n in historial:
        mes = n['fecha'][:7]
        por_mes[mes] = por_mes.get(mes, 0) + n['total']
        noches_mes[mes] = noches_mes.get(mes, 0) + 1
    meses = sorted(por_mes.keys())

    ul = historial[-1]
    pc3 = ul.get('por_caja', {})
    cj  = [pc3.get(i, pc3.get(str(i), 0)) for i in [1,2,3]]
    if not any(cj):
        ops_cj = [{},{},[0,0,0]][0]
        for t in ul.get('transactions', []):
            k = int(t.get('caja', 1)) - 1
            if 0 <= k <= 2: cj[k] = cj[k] + t['amount']

    ops_cj = [0, 0, 0]
    for t in ul.get('transactions', []):
        k = int(t.get('caja', 1)) - 1
        if 0 <= k <= 2: ops_cj[k] += 1

    medal_color = {1: GOLD, 2: '9E9E9E', 3: 'CD7F32'}

    # ════════════════════════════════════════════════════════
    # HOJA 1 — RESUMEN GENERAL
    # ════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Resumen'
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = GOLD
    ws1.freeze_panes = 'A4'

    title_row(ws1, 1,
        'JAGGER CLUB — RESUMEN GENERAL',
        f'Exportado el {datetime.now().strftime("%d/%m/%Y %H:%M")}  ·  {len(historial)} noche(s) registrada(s)')

    # KPIs fila 3 (3 pares lado a lado)
    ws1.row_dimensions[3].height = 6
    kpi_block(ws1, 4, 1, 'TOTAL FACTURADO',    total_all,   '#.##0')
    kpi_block(ws1, 4, 3, 'NOCHES',             len(historial), 'General', BLUE)
    kpi_block(ws1, 4, 5, 'OPERACIONES',        total_ops,   'General', BLUE)
    kpi_block(ws1, 5, 1, 'PROMEDIO / NOCHE',   avg_noche,   '#.##0')
    kpi_block(ws1, 5, 3, 'PROM / OPERACIÓN',   total_all/total_ops if total_ops else 0, '#.##0')
    kpi_block(ws1, 5, 5, 'MEJOR NOCHE',        mejor,       '#.##0')
    ws1.row_dimensions[6].height = 8

    # Ranking general
    section(ws1, 7, '  RANKING GENERAL — TOP 20', 1, 5)
    header_row(ws1, 8, ['#', 'CLIENTE', 'TOTAL ($)', '% DEL TOTAL', 'NOCHES'], 1, GOLD)
    for i, (nm, tot) in enumerate(rk_sorted[:20], 1):
        alt = (i % 2 == 0)
        mc = medal_color.get(i)
        data_cell(ws1, 8+i, 1, f'#{i}', bold=i<=3, alt=alt, h='center', color=mc)
        data_cell(ws1, 8+i, 2, nm,      bold=i<=3, alt=alt, h='left',   color=mc)
        money(ws1, 8+i, 3, tot, alt=alt, bold=i<=3, color=mc)
        pct(ws1, 8+i, 4, tot/total_all if total_all else 0, alt=alt)
        data_cell(ws1, 8+i, 5, noches_cliente.get(nm, 0), alt=alt, h='center')

    base = 10 + len(rk_sorted[:20])
    ws1.row_dimensions[base].height = 10

    # Por mes
    section(ws1, base+1, '  FACTURACIÓN POR MES', 1, 5)
    header_row(ws1, base+2, ['MES', 'TOTAL ($)', '% DEL TOTAL', 'NOCHES', 'PROM/NOCHE'], 1, GOLD)
    ms = base + 3
    for i, mes in enumerate(meses):
        alt = (i%2==0)
        data_cell(ws1, ms+i, 1, mes, alt=alt)
        money(ws1, ms+i, 2, por_mes[mes], alt=alt)
        pct(ws1, ms+i, 3, por_mes[mes]/total_all if total_all else 0, alt=alt)
        data_cell(ws1, ms+i, 4, noches_mes.get(mes,0), alt=alt, h='center')
        money(ws1, ms+i, 5, por_mes[mes]/noches_mes[mes] if noches_mes.get(mes) else 0, alt=alt)
    ms_end = ms + len(meses) - 1

    # Evolución noche a noche — tabla completa
    ev = ms_end + 3
    ws1.row_dimensions[ev].height = 10
    section(ws1, ev+1, '  EVOLUCIÓN NOCHE A NOCHE', 1, 5)
    header_row(ws1, ev+2, ['FECHA', 'TOTAL ($)', 'OPS', 'PROM/OP ($)', 'CAJA TOP'], 1, GOLD)
    ev_d = ev + 3
    for i, n in enumerate(historial):
        alt = (i%2==0)
        data_cell(ws1, ev_d+i, 1, n['fecha'], alt=alt)
        money(ws1, ev_d+i, 2, n['total'], alt=alt)
        data_cell(ws1, ev_d+i, 3, n['operaciones'], alt=alt, h='center')
        money(ws1, ev_d+i, 4, n['total']/n['operaciones'] if n['operaciones'] else 0, alt=alt)
        pc = n.get('por_caja', {}); top = max(pc, key=lambda k: pc[k], default='—')
        data_cell(ws1, ev_d+i, 5, _caja_nombre_hist(n, top) if str(top).lstrip('-').isdigit() else str(top), alt=alt)
    ev_end = ev_d + len(historial) - 1

    # ── Gráfico evolución — UNA línea, un punto por noche, colores distintos por marcador ──
    # El gráfico va DEBAJO de todas las tablas
    chart_anchor_row = ev_end + 3
    ws1.row_dimensions[chart_anchor_row].height = 8

    EVOL_COLORS = ['C9A227','1A4A8A','1E6B3C','A83030','7B4EA0',
                   '4A6B8A','C87320','1E5A4A','6B1A3A','2A3A6B',
                   '3A6A9A','4A9A3A','C94A27','9A27C9','6B3A1A']

    if len(historial) >= 1:
        # Una sola serie = la línea que sube/baja
        ch_l = LineChart()
        ch_l.title = 'Evolución noche a noche'
        ch_l.style = 10
        ch_l.width = 28; ch_l.height = 14
        ch_l.y_axis.numFmt = '#,##0'
        ch_l.y_axis.title = 'Total ($)'
        ch_l.x_axis.title = 'Noche'

        # Datos: todos los totales en una sola serie
        ch_l.add_data(Reference(ws1, min_col=2, min_row=ev+2, max_row=ev_end),
                      titles_from_data=True)
        ch_l.set_categories(Reference(ws1, min_col=1, min_row=ev_d, max_row=ev_end))

        # Línea dorada
        s = ch_l.series[0]
        s.graphicalProperties.line.solidFill = GOLD
        s.graphicalProperties.line.width = 22000

        # Marcadores — color diferente por punto usando DataPoint
        s.marker.symbol = 'circle'
        s.marker.size = 10
        s.marker.graphicalProperties.solidFill = GOLD
        s.marker.graphicalProperties.ln.solidFill = BLACK

        for i in range(len(historial)):
            pt = DataPoint(idx=i)
            c = EVOL_COLORS[i % len(EVOL_COLORS)]
            pt.marker = s.marker.__class__()
            pt.marker.symbol = 'circle'
            pt.marker.size = 10
            pt.marker.graphicalProperties.solidFill = c
            pt.marker.graphicalProperties.ln.solidFill = BLACK
            s.dPt.append(pt)

        ws1.add_chart(ch_l, f'A{chart_anchor_row}')

    # Anchos generosos — elimina todos los ####
    for c, w in enumerate([18, 16, 18, 16, 18, 16, 1, 18, 18, 18, 18, 18], 1):
        col_w(ws1, c, w)

    # ════════════════════════════════════════════════════════
    # HOJA 2 — HISTORIAL DE NOCHES
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Historial')
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = DARK
    ws2.freeze_panes = 'A4'

    title_row(ws2, 1, 'HISTORIAL DE NOCHES',
        f'{len(historial)} noche(s)  ·  Acumulado: ${total_all:,.0f}'.replace(',','.'))

    # Cajas presentes en TODO el historial → columnas dinámicas
    caja_ids_hist = []
    for n in historial:
        for k in (n.get('por_caja', {}) or {}):
            try: cid = int(k)
            except (ValueError, TypeError): continue
            if cid not in caja_ids_hist: caja_ids_hist.append(cid)
    caja_ids_hist.sort()
    def _nombre_caja_global(cid):
        for nn in reversed(historial):
            snap = nn.get('cajas_nombres', {}) or {}
            if str(cid) in snap: return snap[str(cid)]
        return _caja_nombre_map().get(cid, 'Caja ' + str(cid))

    cols2 = ['FECHA', 'HORA CIERRE', 'TOTAL ($)', 'OPS', 'PROM/OP ($)', '1° LUGAR', '2° LUGAR', '3° LUGAR'] \
        + [f'{_nombre_caja_global(cid).upper()} ($)' for cid in caja_ids_hist]
    header_row(ws2, 3, cols2, 1, DARK)
    # fix: re-apply white font since DARK bg
    for ci in range(1, len(cols2)+1):
        ws2.cell(row=3, column=ci).font = Font(bold=True, color=WHITE, size=9, name='Calibri')

    for idx, n in enumerate(reversed(historial)):
        r = 4 + idx; alt = (idx%2==0)
        pc = n.get('por_caja', {})
        rk = sorted(n.get('ranking', []), key=lambda x: -x['total'])
        def top(pos): return rk[pos]['name'] if len(rk) > pos else '—'
        vals = [
            n['fecha'], n.get('hora_cierre','—'), n['total'], n['operaciones'],
            n['total']/n['operaciones'] if n['operaciones'] else 0,
            top(0), top(1), top(2),
        ] + [pc.get(cid, pc.get(str(cid), 0)) for cid in caja_ids_hist]
        fmts = [None, None, '#.##0', 'General', '#.##0', None, None, None] + ['#.##0'] * len(caja_ids_hist)
        cols_color = [None, None, GREEN, BLUE, GREEN, medal_color.get(1), medal_color.get(2), medal_color.get(3)] \
            + [GREEN] * len(caja_ids_hist)
        for ci, (v, fmt, cc) in enumerate(zip(vals, fmts, cols_color), 1):
            bg = LGRAY if alt else WHITE
            c2 = ws2.cell(row=r, column=ci, value=v)
            c2.font = Font(bold=ci in [1,3], color=cc or BLACK, size=9, name='Calibri')
            c2.fill = fill(bg); c2.border = BDR
            c2.alignment = align('right' if fmt=='#.##0' else 'center')
            if fmt: c2.number_format = fmt
            row_h(ws2, r, 17)

    for c, w in enumerate([13, 12, 16, 7, 16, 20, 20, 20] + [16] * len(caja_ids_hist), 1):
        col_w(ws2, c, w)

    # ════════════════════════════════════════════════════════
    # HOJAS POR NOCHE — una hoja por cada noche del mes actual
    # ════════════════════════════════════════════════════════
    mes_actual = datetime.now().strftime('%Y-%m')
    noches_mes = [n for n in historial if n['fecha'].startswith(mes_actual)]
    if not noches_mes:
        noches_mes = historial  # fallback: todas las noches

    def make_noche_sheet(n, sheet_name):
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = GOLD
        ws.freeze_panes = 'A5'

        # Datos de la noche — cajas dinámicas
        pc = {int(k): v for k, v in (n.get('por_caja', {}) or {}).items()}
        txs_night = n.get('transactions', [])
        t_n  = n['total']; ops_n = n['operaciones']
        ops_caja, rec_caja = {}, {}
        for t in txs_night:
            c = int(t.get('caja', 1))
            ops_caja[c] = ops_caja.get(c, 0) + 1
            rec_caja[c] = rec_caja.get(c, 0) + t.get('amount', 0)
        ids_n = sorted(set(pc.keys()) | set(ops_caja.keys()))
        tot_caja = {cid: (pc[cid] if cid in pc else rec_caja.get(cid, 0)) for cid in ids_n}
        CAJA_XLS = ['1A4A8A','1E6B3C','C9A227','A83030','7B4EA0','C87320','1E5A4A','2A3A6B','4A9A3A','9A27C9']

        title_row(ws, 1, f'DETALLE — {n["fecha"]}',
            f'Cierre: {n.get("hora_cierre","—")}  ·  Total: ${t_n:,.0f}  ·  {ops_n} ops  ·  Prom/op: ${t_n/ops_n:,.0f}'.replace(',','.'))
        ws.row_dimensions[3].height = 6

        # KPIs
        kpi_block(ws, 4, 1, 'TOTAL NOCHE',   t_n,   '#.##0')
        kpi_block(ws, 4, 3, 'OPERACIONES',    ops_n, 'General', BLUE)
        kpi_block(ws, 4, 5, 'PROMEDIO / OP',  t_n/ops_n if ops_n else 0, '#.##0')
        kpi_block(ws, 5, 1, 'CLIENTES',       len(n.get('ranking', [])), 'General', BLUE)
        kpi_block(ws, 5, 3, 'N° DE CAJAS',    len([c for c in ids_n if tot_caja.get(c, 0) > 0]), 'General', BLUE)
        kpi_block(ws, 5, 5, 'MAYOR CAJA',     max(tot_caja.values()) if tot_caja else 0, '#.##0')
        ws.row_dimensions[6].height = 8

        # ── Ranking de la noche ──────────────────────────────
        section(ws, 7, '  RANKING DE LA NOCHE', 1, 5)
        header_row(ws, 8, ['#', 'CLIENTE', 'MESA', 'TOTAL ($)', '% DEL TOTAL'], 1, GOLD)
        rk_n = n.get('ranking', [])
        for i, rv in enumerate(rk_n, 1):
            alt = (i%2==0); mc = medal_color.get(i)
            data_cell(ws, 8+i, 1, f'#{i}',           bold=i<=3, alt=alt, h='center', color=mc)
            data_cell(ws, 8+i, 2, rv['name'],         bold=i<=3, alt=alt, h='left',   color=mc)
            data_cell(ws, 8+i, 3, rv.get('mesa','—'), alt=alt, h='center')
            money(ws, 8+i, 4, rv['total'], alt=alt, bold=i<=3, color=mc)
            pct(ws, 8+i, 5, rv['total']/t_n if t_n else 0, alt=alt)

        # ── Distribución por caja — con pie chart bien posicionado ──
        dist_start = 10 + len(rk_n)
        ws.row_dimensions[dist_start].height = 8
        section(ws, dist_start+1, '  DISTRIBUCIÓN POR CAJA', 1, 5)
        header_row(ws, dist_start+2, ['CAJA', 'TOTAL ($)', '% DEL TOTAL', 'OPS', 'PROM/OP ($)'], 1, GOLD)
        cs = dist_start + 3
        for i, cid in enumerate(ids_n):
            alt = (i%2==0); col = CAJA_XLS[i % len(CAJA_XLS)]
            tot = tot_caja.get(cid, 0); ops = ops_caja.get(cid, 0)
            data_cell(ws, cs+i, 1, _caja_nombre_hist(n, cid), bold=True, alt=alt, h='left', color=col)
            money(ws, cs+i, 2, tot, alt=alt, color=col)
            pct(ws, cs+i, 3, tot/t_n if t_n else 0, alt=alt)
            data_cell(ws, cs+i, 4, ops, alt=alt, h='center')
            money(ws, cs+i, 5, tot/ops if ops else 0, alt=alt)
        n_cajas = len(ids_n)

        # ── Todas las operaciones ────────────────────────────
        op_start = cs + max(1, n_cajas) + 2
        ws.row_dimensions[op_start].height = 8
        section(ws, op_start+1, '  TODAS LAS OPERACIONES', 1, 6)
        header_row(ws, op_start+2, ['HORA', 'CLIENTE', 'MESA', 'MONTO ($)', 'CAJA', 'PRODUCTOS'], 1, DARK)
        for ci in range(1, 7):
            ws.cell(row=op_start+2, column=ci).font = Font(bold=True, color=WHITE, size=9, name='Calibri')
        txs_n = n.get('transactions', [])
        for i, t in enumerate(txs_n):
            alt = (i%2==0)
            data_cell(ws, op_start+3+i, 1, t.get('time','—'), alt=alt)
            data_cell(ws, op_start+3+i, 2, t['name'], bold=True, alt=alt, h='left', color=BLUE)
            data_cell(ws, op_start+3+i, 3, t.get('mesa','—'), alt=alt)
            money(ws, op_start+3+i, 4, t['amount'], alt=alt)
            data_cell(ws, op_start+3+i, 5, _caja_nombre_hist(n, t.get('caja', 1)), alt=alt)
            istr = ' · '.join(f"{it.get('cantidad',it.get('qty',1))}x {it['nombre']}" for it in t.get('items',[])) or '—'
            cc2 = ws.cell(row=op_start+3+i, column=6, value=istr)
            cc2.font = Font(color=DGRAY, size=9, name='Calibri')
            cc2.fill = fill(LGRAY if alt else WHITE); cc2.border = BDR
            cc2.alignment = align('left', indent=1); row_h(ws, op_start+3+i, 17)

        # Pie chart — debajo de TODAS las tablas y operaciones
        pie_anchor = op_start + 3 + len(txs_n) + 3
        if n_cajas and any(tot_caja.values()):
            ch_pie = PieChart(); ch_pie.title = 'Distribución por caja'; ch_pie.style = 10
            ch_pie.width = 18; ch_pie.height = 12
            ch_pie.add_data(Reference(ws, min_col=2, min_row=dist_start+2, max_row=cs+n_cajas-1), titles_from_data=True)
            ch_pie.set_categories(Reference(ws, min_col=1, min_row=cs, max_row=cs+n_cajas-1))
            for idx in range(n_cajas):
                pt = DataPoint(idx=idx)
                pt.graphicalProperties.solidFill = CAJA_XLS[idx % len(CAJA_XLS)]
                ch_pie.series[0].dPt.append(pt)
            ws.add_chart(ch_pie, f'A{pie_anchor}')

        # Anchos — generosos para evitar palabras cortadas y ####
        for c, w in enumerate([7, 24, 8, 16, 14, 44, 1, 18, 18, 18, 18, 18], 1):
            col_w(ws, c, w)
        # Columnas KPI (1-6): label + value pairs
        for c, w in enumerate([18, 16, 18, 16, 18, 16], 1):
            col_w(ws, c, w)

    # Generar una hoja por cada noche del mes (más reciente primero)
    for n in reversed(noches_mes):
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(n['fecha'], '%Y-%m-%d')
            sname = d.strftime('%d-%m')
        except Exception:
            sname = n['fecha'][-5:].replace('-','/')
        make_noche_sheet(n, sname)

    # ════════════════════════════════════════════════════════
    # HOJA — CLIENTES (siempre al final)
    # ════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Clientes')
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = BLUE
    ws4.freeze_panes = 'A4'

    title_row(ws4, 1, 'RANKING DE CLIENTES — HISTÓRICO',
        f'{len(rk_sorted)} clientes únicos  ·  Total facturado: ${total_all:,.0f}'.replace(',','.'))

    ws4.row_dimensions[3].height = 6
    header_row(ws4, 3, ['#', 'CLIENTE', 'TOTAL HISTÓRICO ($)', '% DEL TOTAL', 'NOCHES', 'PROM/NOCHE ($)'], 1, DARK)
    for ci in range(1, 7):
        ws4.cell(row=3, column=ci).font = Font(bold=True, color=WHITE, size=9, name='Calibri')
    for i, (nm, tot) in enumerate(rk_sorted, 1):
        alt = (i%2==0); mc = medal_color.get(i); nn = noches_cliente.get(nm, 1)
        data_cell(ws4, 3+i, 1, f'#{i}', bold=i<=3, alt=alt, h='center', color=mc)
        data_cell(ws4, 3+i, 2, nm,      bold=i<=3, alt=alt, h='left',   color=mc)
        money(ws4, 3+i, 3, tot, alt=alt, bold=i<=3, color=mc)
        pct(ws4, 3+i, 4, tot/total_all if total_all else 0, alt=alt)
        data_cell(ws4, 3+i, 5, nn, alt=alt, h='center')
        money(ws4, 3+i, 6, tot/nn, alt=alt)
    for c, w in enumerate([6, 24, 18, 12, 8, 14], 1): col_w(ws4, c, w)

    # ── Generar y enviar ──────────────────────────────────────
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import send_file
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf, as_attachment=True,
                     download_name=f'jagger_vip_{fecha_str}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
@app.route('/api/tarjetas/recargar', methods=['POST'])
def recargar_tarjeta():
    body = request.get_json() or {}
    codigo = str(body.get('codigo', ''))
    try:
        monto = float(body.get('monto', 0))
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'Monto inválido'}), 400
    if not codigo or monto <= 0:
        return jsonify({'ok': False, 'error': 'Datos inválidos'}), 400
    with lock:
        data = load_data()
        if codigo not in data.get('tarjetas', {}):
            return jsonify({'ok': False, 'error': 'Tarjeta no encontrada'}), 404
        import time as _time
        recarga_id = str(int(_time.time()*1000))
        data['tarjetas'][codigo]['saldo_actual'] += monto
        nuevo_saldo = data['tarjetas'][codigo]['saldo_actual']
        if 'recargas' not in data['tarjetas'][codigo]:
            data['tarjetas'][codigo]['recargas'] = []
        from datetime import datetime as _dt
        data['tarjetas'][codigo]['recargas'].append({
            'id': recarga_id, 'monto': monto,
            'hora': _dt.now().strftime('%H:%M'), 'fecha': _dt.now().strftime('%d/%m')
        })
        save_data(data)
    return jsonify({'ok': True, 'nuevo_saldo': nuevo_saldo, 'recarga_id': recarga_id})

@app.route('/api/tarjetas/recarga/<codigo>/<recarga_id>', methods=['DELETE'])
def eliminar_recarga(codigo, recarga_id):
    with lock:
        data = load_data()
        t = data.get('tarjetas', {}).get(codigo)
        if not t:
            return jsonify({'ok': False, 'error': 'Tarjeta no encontrada'}), 404
        recargas = t.get('recargas', [])
        rec = next((r for r in recargas if r['id'] == recarga_id), None)
        if not rec:
            return jsonify({'ok': False, 'error': 'Recarga no encontrada'}), 404
        t['saldo_actual'] = max(0, t['saldo_actual'] - rec['monto'])
        t['recargas'] = [r for r in recargas if r['id'] != recarga_id]
        nuevo_saldo = t['saldo_actual']
        save_data(data)
    return jsonify({'ok': True, 'nuevo_saldo': nuevo_saldo})

@app.route('/api/tarjetas/recarga/<codigo>/<recarga_id>', methods=['PUT'])
def editar_recarga(codigo, recarga_id):
    body = request.get_json() or {}
    try:
        nuevo_monto = float(body.get('monto', 0))
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'Monto inválido'}), 400
    if nuevo_monto <= 0:
        return jsonify({'ok': False, 'error': 'Monto inválido'}), 400
    with lock:
        data = load_data()
        t = data.get('tarjetas', {}).get(codigo)
        if not t:
            return jsonify({'ok': False, 'error': 'Tarjeta no encontrada'}), 404
        recargas = t.get('recargas', [])
        rec = next((r for r in recargas if r['id'] == recarga_id), None)
        if not rec:
            return jsonify({'ok': False, 'error': 'Recarga no encontrada'}), 404
        diff = nuevo_monto - rec['monto']
        t['saldo_actual'] = max(0, t['saldo_actual'] + diff)
        rec['monto'] = nuevo_monto
        nuevo_saldo = t['saldo_actual']
        save_data(data)
    return jsonify({'ok': True, 'nuevo_saldo': nuevo_saldo})

@app.route('/api/auth', methods=['POST'])
def auth():
    body = request.get_json() or {}
    cfg = load_config()
    if str(body.get('password', '')) == str(cfg.get('password', '1212')):
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'Contraseña incorrecta'}), 401

@app.route('/api/auth/change', methods=['POST'])
def auth_change():
    body = request.get_json() or {}
    cfg = load_config()
    if str(body.get('current', '')) != str(cfg.get('password', '1212')):
        return jsonify({'ok': False, 'error': 'PIN actual incorrecto'}), 401
    nueva = str(body.get('new', '')).strip()
    if not nueva.isdigit() or len(nueva) != 4:
        return jsonify({'ok': False, 'error': 'El PIN debe tener exactamente 4 dígitos numéricos'}), 400
    cfg['password'] = nueva
    save_config(cfg)
    return jsonify({'ok': True})

# ── Login / Logout ──────────────────────────────────────────────────────────

@app.route('/api/login', methods=['POST'])
def api_login():
    body = request.get_json() or {}
    role = str(body.get('role', ''))
    pin = str(body.get('pin', ''))
    cfg = load_config()
    # Compatibilidad con los nombres de rol viejos → est<id>
    LEGACY = {'cajaabajo': 'est1', 'cajaextendido': 'est2', 'cajavip': 'est3', 'tarjetas': 'est4'}
    if role in LEGACY:
        role = LEGACY[role]
    # Determinar el PIN esperado según el rol
    if role == 'manager':
        pin_ok = str(cfg.get('pin_manager', '4321'))
    elif role.startswith('est'):
        e = _estacion(role[3:], cfg)
        if not e:
            return jsonify({'ok': False, 'error': 'Estación inválida'}), 400
        pin_ok = str(e.get('pin', '1234'))
    else:
        return jsonify({'ok': False, 'error': 'Rol inválido'}), 400
    if pin == pin_ok:
        session.permanent = True
        session[role + '_auth'] = True
        session['session_version'] = cfg.get('session_version', 0)
        if role == 'manager':
            # El manager destraba todas las estaciones
            for e in _estaciones(cfg):
                session['est' + str(e['id']) + '_auth'] = True
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'PIN incorrecto'}), 401

@app.route('/api/session', methods=['GET'])
def get_session_info():
    """Qué estaciones tiene abiertas esta sesión.
    El manager destraba todas las demás, así que se refleja acá."""
    cfg = load_config()
    out = {'manager': check_auth('manager')}
    for e in _estaciones(cfg):
        out['est' + str(e['id'])] = check_auth('est' + str(e['id']))
    # Alias de compatibilidad con lectores viejos
    out['cajaabajo']     = out.get('est1', False)
    out['cajaextendido'] = out.get('est2', False)
    out['cajavip']       = out.get('est3', False)
    out['tarjetas']      = out.get('est4', False)
    return jsonify(out)

@app.route('/api/logout', methods=['POST'])
def api_logout():
    body = request.get_json() or {}
    role = body.get('role')
    if role:
        session.pop(role + '_auth', None)
    else:
        session.clear()
    return jsonify({'ok': True})

# ── Menu CRUD ───────────────────────────────────────────────────────────────

DEFAULT_MENU = [
    # CHAMPAGNE
    {"categoria":"Champagne","nombre":"Norton Cosecha Tardia","precio":50000},
    {"categoria":"Champagne","nombre":"Mumm","precio":60000},
    {"categoria":"Champagne","nombre":"Chandon","precio":80000},
    {"categoria":"Champagne","nombre":"Baron B","precio":100000},
    # BOTTLE SERVICE
    {"categoria":"Bottle Service","nombre":"Sernova","precio":85000},
    {"categoria":"Bottle Service","nombre":"Absolut","precio":105000},
    {"categoria":"Bottle Service","nombre":"Red Label","precio":120000},
    {"categoria":"Bottle Service","nombre":"Red Label Litro","precio":140000},
    {"categoria":"Bottle Service","nombre":"Black Label","precio":170000},
    {"categoria":"Bottle Service","nombre":"Jagermeister","precio":120000},
    {"categoria":"Bottle Service","nombre":"Beefeater","precio":120000},
    {"categoria":"Bottle Service","nombre":"Gin Blu","precio":100000},
    {"categoria":"Bottle Service","nombre":"Malibu","precio":100000},
    {"categoria":"Bottle Service","nombre":"Fernet Branca","precio":90000},
    {"categoria":"Bottle Service","nombre":"Ramazotti","precio":90000},
    # IMPORTADOS
    {"categoria":"Importados","nombre":"Belvedere Luminous","precio":180000},
    {"categoria":"Importados","nombre":"Belvedere Luminous X2","precio":340000},
    {"categoria":"Importados","nombre":"Cliquot","precio":220000},
    {"categoria":"Importados","nombre":"Moet Nectar","precio":240000},
    {"categoria":"Importados","nombre":"Moet Ice","precio":250000},
    {"categoria":"Importados","nombre":"Nuvo","precio":180000},
    {"categoria":"Importados","nombre":"Nuvo X2","precio":340000},
    # BEBIDAS
    {"categoria":"Bebidas","nombre":"Speed","precio":7000},
    {"categoria":"Bebidas","nombre":"Speed X6","precio":35000},
    {"categoria":"Bebidas","nombre":"Jugo","precio":7000},
    {"categoria":"Bebidas","nombre":"Lata Coca","precio":6500},
    {"categoria":"Bebidas","nombre":"Lata Ton","precio":7000},
    {"categoria":"Bebidas","nombre":"Agua","precio":5500},
    {"categoria":"Bebidas","nombre":"Budweiser","precio":7000},
    {"categoria":"Bebidas","nombre":"Budweiser X6","precio":35000},
    # TRAGOS
    {"categoria":"Tragos","nombre":"Sernova","precio":8000},
    {"categoria":"Tragos","nombre":"Absolut","precio":9500},
    {"categoria":"Tragos","nombre":"Hodlmoser","precio":9000},
    {"categoria":"Tragos","nombre":"Fernet","precio":8000},
    {"categoria":"Tragos","nombre":"Gin","precio":9500},
    {"categoria":"Tragos","nombre":"Campari","precio":8000},
    {"categoria":"Tragos","nombre":"Malibu","precio":8500},
    # SHOTS
    {"categoria":"Shots","nombre":"Hodlmoser","precio":7000},
    {"categoria":"Shots","nombre":"Absolut","precio":7000},
]

def ensure_menu():
    """Populate default menu into _db if it's empty."""
    with lock:
        data = load_data()
        if not data.get('menu'):
            for i, item in enumerate(DEFAULT_MENU, 1):
                data['menu'].append({'id': i, 'categoria': item['categoria'], 'nombre': item['nombre'], 'precio': float(item['precio'])})
            data['menu_id_counter'] = len(DEFAULT_MENU)
            save_data(data)

@app.route('/api/menu/reset', methods=['POST'])
def reset_menu():
    with lock:
        data = load_data()
        data['menu'] = []
        for i, item in enumerate(DEFAULT_MENU, 1):
            data['menu'].append({'id':i,'categoria':item['categoria'],'nombre':item['nombre'],'precio':float(item['precio'])})
        data['menu_id_counter'] = len(DEFAULT_MENU)
        save_data(data)
        bump_menu()
    return jsonify({'ok': True, 'count': len(DEFAULT_MENU)})

@app.route('/api/menu', methods=['GET'])
def get_menu():
    ensure_menu()
    with lock:
        data = load_data()
    return jsonify(data.get('menu', []))

def _categorias(data):
    """Lista de categorías: las de los productos + las creadas a mano, ordenadas."""
    delmenu = {str(p.get('categoria', '')).strip() for p in data.get('menu', []) if p.get('categoria')}
    extras = {str(c).strip() for c in data.get('categorias', []) if str(c).strip()}
    return sorted(delmenu | extras, key=str.lower)

@app.route('/api/categorias', methods=['GET'])
def get_categorias():
    ensure_menu()
    with lock:
        data = load_data()
        cats = _categorias(data)
    return jsonify(cats)

@app.route('/api/categorias', methods=['POST'])
def add_categoria():
    body = request.get_json() or {}
    nombre = str(body.get('nombre', '')).strip()
    if not nombre:
        return jsonify({'ok': False, 'error': 'Nombre vacío'}), 400
    with lock:
        data = load_data()
        if 'categorias' not in data:
            data['categorias'] = []
        existentes = {c.lower() for c in _categorias(data)}
        if nombre.lower() not in existentes:
            data['categorias'].append(nombre)
            save_data(data)
            bump_menu()
        cats = _categorias(data)
    return jsonify({'ok': True, 'categorias': cats})

@app.route('/api/categorias/<nombre>', methods=['DELETE'])
def del_categoria(nombre):
    """Borra una categoría solo si está vacía (sin productos)."""
    with lock:
        data = load_data()
        con_productos = sum(1 for p in data.get('menu', []) if str(p.get('categoria', '')) == nombre)
        if con_productos:
            return jsonify({'ok': False, 'error': f'La categoría tiene {con_productos} producto(s). Vaciala primero.'}), 400
        data['categorias'] = [c for c in data.get('categorias', []) if c != nombre]
        save_data(data)
        bump_menu()
        cats = _categorias(data)
    return jsonify({'ok': True, 'categorias': cats})

@app.route('/api/menu', methods=['POST'])
def add_menu_item():
    body = request.get_json() or {}
    if not body.get('nombre') or body.get('precio') is None:
        return jsonify({'ok': False, 'error': 'Faltan campos'}), 400
    with lock:
        data = load_data()
        if 'menu' not in data: data['menu'] = []
        counter = data.get('menu_id_counter', len(data['menu']))
        counter += 1
        item = {
            'id': counter,
            'categoria': str(body.get('categoria', 'Otros')),
            'nombre': str(body['nombre']),
            'precio': float(body['precio']),
        }
        data['menu'].append(item)
        data['menu_id_counter'] = counter
        save_data(data)
        bump_menu()
    return jsonify({'ok': True, 'item': item})

@app.route('/api/menu/<int:mid>', methods=['PUT'])
def update_menu_item(mid):
    body = request.get_json() or {}
    with lock:
        data = load_data()
        item = next((x for x in data.get('menu', []) if x['id'] == mid), None)
        if not item:
            return jsonify({'ok': False, 'error': 'No encontrado'}), 404
        if 'nombre' in body: item['nombre'] = str(body['nombre'])
        if 'precio' in body: item['precio'] = float(body['precio'])
        if 'categoria' in body: item['categoria'] = str(body['categoria'])
        save_data(data)
        bump_menu()
    return jsonify({'ok': True})

@app.route('/api/menu/<int:mid>', methods=['DELETE'])
def delete_menu_item(mid):
    with lock:
        data = load_data()
        data['menu'] = [x for x in data.get('menu', []) if x['id'] != mid]
        save_data(data)
        bump_menu()
    return jsonify({'ok': True})

# ── Publicidad ──────────────────────────────────────────────────────────────

import werkzeug.utils as _wu

@app.route('/api/publicidad/estado', methods=['GET'])
def pub_estado():
    import time as _t
    return jsonify({
        'mostrar_ts': _state.get('publicidad_mostrar_ts', 0),
        'activa': _state['publicidad_activa'],
        'url': _state['publicidad_url'],
        'frecuencia': _state['publicidad_frecuencia'],
        'mostrar_ts': _state.get('publicidad_mostrar_ts', 0),
        'server_time': _t.time(),
    })

@app.route('/api/publicidad/activar', methods=['POST'])
def pub_activar():
    body = request.get_json() or {}
    with lock:
        _state['publicidad_activa'] = True
        _state['publicidad_url'] = str(body.get('url', _state['publicidad_url']))
        _state['publicidad_frecuencia'] = int(body.get('frecuencia', _state['publicidad_frecuencia']))
        save_state()
    return jsonify({'ok': True})

@app.route('/api/publicidad/desactivar', methods=['POST'])
def pub_desactivar():
    with lock:
        _state['publicidad_activa'] = False
        save_state()
    return jsonify({'ok': True})

@app.route('/api/publicidad/mostrar-ahora', methods=['POST'])
def pub_mostrar_ahora():
    import time as _t
    with lock:
        _state['publicidad_mostrar_ts'] = _t.time()
    return jsonify({'ok': True})

@app.route('/api/publicidad/upload', methods=['POST'])
def pub_upload():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No se envió archivo'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.mp4'):
        return jsonify({'ok': False, 'error': 'Solo se aceptan archivos mp4'}), 400
    os.makedirs(VIDEOS_DIR, exist_ok=True)
    fname = _wu.secure_filename(f.filename)
    fpath = os.path.join(VIDEOS_DIR, fname)
    f.save(fpath)
    url = '/api/publicidad/video/' + fname
    with lock:
        _state['publicidad_url'] = url
        save_state()
    return jsonify({'ok': True, 'url': url})

@app.route('/api/publicidad/video/<path:filename>')
def serve_video(filename):
    from flask import send_from_directory
    return send_from_directory(VIDEOS_DIR, filename)

# ── Config / PINs ───────────────────────────────────────────────────────────

@app.route('/api/config/pines', methods=['POST'])
def update_pines():
    """Actualiza el PIN del manager y/o de cada estación.
    body = {pin_manager?: '1234', estaciones?: {<id>: '1234', ...}}"""
    body = request.get_json() or {}
    with lock:
        cfg = load_config()
        changed = False
        if 'pin_manager' in body:
            v = str(body['pin_manager']).strip()
            if not (v.isdigit() and len(v) == 4):
                return jsonify({'ok': False, 'error': 'PIN de manager inválido'}), 400
            cfg['pin_manager'] = v
            changed = True
        pins = body.get('estaciones', {}) or {}
        for e in cfg.get('estaciones', []):
            key = str(e['id'])
            if key in pins:
                v = str(pins[key]).strip()
                if not (v.isdigit() and len(v) == 4):
                    return jsonify({'ok': False, 'error': f'PIN inválido para {e.get("nombre", key)}'}), 400
                e['pin'] = v
                changed = True
        if changed:
            cfg['session_version'] = cfg.get('session_version', 0) + 1
            save_config(cfg)
    return jsonify({'ok': True})

@app.route('/api/config/pines', methods=['GET'])
def get_pines():
    cfg = load_config()
    return jsonify({
        'pin_manager': cfg.get('pin_manager', ''),
        'estaciones': [{'id': e['id'], 'tipo': e.get('tipo', 'caja'), 'nombre': e.get('nombre', ''), 'pin': e.get('pin', '')}
                       for e in _estaciones(cfg)],
    })

# ── Design state ─────────────────────────────────────────────────────────────

@app.route('/api/design', methods=['GET'])
def get_design():
    return jsonify(_state.get('design', {}))

@app.route('/api/design', methods=['POST'])
def set_design():
    body = request.get_json() or {}
    with lock:
        if 'design' not in _state:
            _state['design'] = {}
        allowed = {'tema','colores','logo','vip','tagline','tagline_color','tagline_glow',
                   'tagline_font','winner_msg','winner_sub','premio','premio_size',
                   'hora_fin','clock_size','deco_activa','petals_activos',
                   'falling_gloves','pink_modo','efecto','tipo_particula','mostrar12','arg_estrellas','arg_fondo','arg_estrellas_on',
                   # ── Sistema de diseño JAGGER 2.0 ──
                   'font',            # Space Grotesk | Arial | Oswald | Montserrat | Bebas Neue | Anton | Archivo Black | Teko
                   'accent',          # nombre del swatch (Lima, Dorado, ...)
                   'acColor',         # hex del acento
                   'efectos',         # Brillos | Flotación | Brillos + Flotación | Estelar | Ninguno
                   'fondo',           # Acento | Dorado | Rosa | Cian | Violeta | Azul | Sin fondo
                   'fondoIntensidad', # Sutil | Visible | Intenso
                   'flot',            # Off | Suave | Media | Fuerte
                   'modo_caja',       # dark | light  (tema de la Caja POS)
                   'premio_on',       # mostrar/ocultar el cartel de premio
                   'textos',          # overrides de copy por pantalla (texto, color, tamaño, on/off)
                   'extras',          # líneas de texto extra agregadas desde el manager
                   'fx',              # {confetti, grilla, animPos, sheen}
                   'relojSize',       # tamaño del reloj en la pantalla
                   # ── Color + brillo por texto (título, ganador, premio) ──
                   'logo_color', 'logo_glow',
                   'vip_color', 'vip_glow',
                   'winner_msg_color', 'winner_msg_glow',
                   'winner_sub_color', 'winner_sub_glow',
                   'premio_color', 'premio_glow',
                   # ── Evento / barra de objetivo del ranking ──
                   'evento_on',        # mostrar la barra de objetivo
                   'evento_premio',    # lo que se gana al llegar (ej: "30% OFF en tragos")
                   'evento_objetivo',  # monto objetivo en $
                   'evento_emojis',    # emojis que decoran la barra
                   'evento_color',     # color del relleno de la barra
                   'evento_msg'}       # mensaje al lograr el objetivo (animación de 1 min)
        updated = False
        for k, v in body.items():
            if k in allowed:
                # No sobreescribir colores con diccionario vacío
                if k == 'colores' and isinstance(v, dict) and len(v) == 0:
                    continue
                _state['design'][k] = v
                updated = True
        if updated:
            import time as _t
            _state['design']['_ts'] = _t.time()
            save_state()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
#  ESTADÍSTICAS Y GANADOR  (secciones nuevas del Panel Manager)
# ══════════════════════════════════════════════════════════════════════════════

def _calcular_ranking():
    """Ranking de la noche en curso: total gastado por persona, desc."""
    data = load_data()
    totals, mesas, codigos = {}, {}, {}
    for t in data.get('transactions', []):
        n = t.get('name', '')
        if not n or t.get('cancelado'):
            continue
        totals[n] = totals.get(n, 0) + t.get('amount', 0)
        if t.get('mesa') and n not in mesas:
            mesas[n] = t['mesa']
        if t.get('tarjeta_codigo') and n not in codigos:
            codigos[n] = t['tarjeta_codigo']
    ranking = [
        {'name': n, 'total': v, 'mesa': mesas.get(n, ''), 'codigo': codigos.get(n, '')}
        for n, v in totals.items()
    ]
    ranking.sort(key=lambda x: -x['total'])
    return ranking


@app.route('/api/ranking', methods=['GET'])
def api_ranking():
    """Ranking en vivo — lo consume la pantalla del ranking y el ganador."""
    with lock:
        ranking = _calcular_ranking()
    return jsonify(ranking)


@app.route('/api/winner/actual', methods=['GET'])
def api_winner_actual():
    """Ganador provisorio: quien más gastó en la noche en curso."""
    with lock:
        ranking = _calcular_ranking()
    if not ranking:
        return jsonify({'hay_ganador': False})
    g = ranking[0]
    return jsonify({
        'hay_ganador': True,
        'name': g['name'],
        'total': g['total'],
        'mesa': g['mesa'],
        'codigo': g['codigo'],
    })


@app.route('/api/estadisticas', methods=['GET'])
def api_estadisticas():
    """Datos para la sección Estadísticas del Panel Manager (cajas dinámicas)."""
    with lock:
        cfg = load_config()
        nombres = _caja_nombre_map(cfg)
        data = load_data()
        # Las operaciones anuladas no suman a ninguna estadística
        txs = [t for t in data.get('transactions', []) if not t.get('cancelado')]

        total = sum(t.get('amount', 0) for t in txs)

        # Ids de caja a reportar: cajas activas (en orden) + las que tengan movimiento
        ids = [int(e['id']) for e in _estaciones(cfg) if e.get('tipo') == 'caja' and e.get('activa', True)]
        for t in txs:
            c = int(t.get('caja', 1))
            if c not in ids:
                ids.append(c)
        cajas = [{'id': cid, 'nombre': nombres.get(cid, 'Caja ' + str(cid))} for cid in ids]

        # Totales y top gastadores por caja
        por_caja = {}
        for cid in ids:
            caja_txs = [t for t in txs if int(t.get('caja', 1)) == cid]
            totales_persona = {}
            for t in caja_txs:
                n = t.get('name', '')
                if n:
                    totales_persona[n] = totales_persona.get(n, 0) + t.get('amount', 0)
            top = sorted(
                [{'name': n, 'total': v} for n, v in totales_persona.items()],
                key=lambda x: -x['total']
            )[:5]
            por_caja[cid] = {
                'nombre': nombres.get(cid, 'Caja ' + str(cid)),
                'total': sum(t.get('amount', 0) for t in caja_txs),
                'operaciones': len(caja_txs),
                'top': top,
            }

        # Consumo por hora, separado por caja (para el gráfico de barras)
        horas = {}
        for t in txs:
            hora = str(t.get('time', ''))[:2]
            if not hora.isdigit():
                continue
            c = int(t.get('caja', 1))
            horas.setdefault(hora, {})
            horas[hora][c] = horas[hora].get(c, 0) + t.get('amount', 0)
        por_hora = [
            {'hora': h + ':00',
             'cajas': {str(cid): v.get(cid, 0) for cid in ids},
             'total': sum(v.values())}
            for h, v in sorted(horas.items())
        ]

        ranking = _calcular_ranking()

    return jsonify({
        'total': total,
        'operaciones': len(txs),
        'cajas': cajas,
        'por_caja': por_caja,
        'por_hora': por_hora,
        'ranking': ranking,
        'ticket_promedio': (total / len(txs)) if txs else 0,
    })


@app.route('/api/noche/resetear', methods=['POST'])
def api_noche_resetear():
    """Limpia el estado de pantalla para arrancar una noche nueva.
    No borra las transacciones (para eso está /api/reset)."""
    with lock:
        _state['winner_show'] = False
        _state['cartel_show'] = False
        _state['cartel_data'] = {}
        save_state()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
#  PANEL DE DUEÑO  /suca  —  branding + pagos (protegido por su propia clave)
# ══════════════════════════════════════════════════════════════════════════════

def check_suca():
    return session.get('suca_auth') == True

def _pago_estado():
    """Calcula el nivel de aviso de pago (0 = ok, 1 = amarillo, 2 = vencido)."""
    p = _state.get('pagos', {}) or {}
    activo = bool(p.get('activo'))
    override = str(p.get('override', 'auto'))
    vence = str(p.get('vence', '') or '')
    dias_restantes = None
    if vence:
        try:
            d = datetime.strptime(vence, '%Y-%m-%d').date()
            dias_restantes = (d - datetime.now().date()).days
        except Exception:
            dias_restantes = None
    nivel = 0
    if activo:
        if override == 'off':
            nivel = 0
        elif override == '1':
            nivel = 1
        elif override == '2':
            nivel = 2
        else:  # 'auto' → por fecha de vencimiento
            if dias_restantes is not None:
                if dias_restantes < 0:
                    nivel = 2
                elif dias_restantes <= int(p.get('dias_aviso', 5) or 0):
                    nivel = 1
    return {
        'nivel': nivel,
        'activo': activo,
        'vence': vence,
        'monto': p.get('monto', 0),
        'nota': p.get('nota', ''),
        'dias_aviso': int(p.get('dias_aviso', 5) or 0),
        'override': override,
        'dias_restantes': dias_restantes,
    }

@app.route('/suca')
def suca_page():
    return render_template('suca.html')

@app.route('/api/suca/session', methods=['GET'])
def suca_session():
    return jsonify({'auth': check_suca()})

@app.route('/api/suca/login', methods=['POST'])
def suca_login():
    body = request.get_json() or {}
    cfg = load_config()
    if str(body.get('password', '')) == str(cfg.get('suca_password', 'suca2026')):
        session.permanent = True
        session['suca_auth'] = True
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'Clave incorrecta'}), 401

@app.route('/api/suca/logout', methods=['POST'])
def suca_logout():
    session.pop('suca_auth', None)
    return jsonify({'ok': True})

@app.route('/api/suca/password', methods=['POST'])
def suca_password_change():
    if not check_suca():
        return jsonify({'ok': False, 'error': 'No autorizado'}), 403
    body = request.get_json() or {}
    nueva = str(body.get('new', '')).strip()
    if len(nueva) < 4:
        return jsonify({'ok': False, 'error': 'La clave debe tener al menos 4 caracteres'}), 400
    cfg = load_config()
    cfg['suca_password'] = nueva
    save_config(cfg)
    return jsonify({'ok': True})

@app.route('/api/branding', methods=['GET'])
def get_branding():
    """Branding vigente — lo consumen todas las pantallas (público)."""
    return jsonify(_state.get('branding', {}))

@app.route('/api/suca/branding', methods=['POST'])
def set_branding():
    if not check_suca():
        return jsonify({'ok': False, 'error': 'No autorizado'}), 403
    body = request.get_json() or {}
    with lock:
        b = _state.setdefault('branding', {})
        for k in ('color', 'logo_url', 'nombre'):
            if k in body:
                b[k] = str(body[k])
        save_state()
    return jsonify({'ok': True, 'branding': _state['branding']})

@app.route('/api/suca/logo', methods=['POST'])
def suca_logo_upload():
    if not check_suca():
        return jsonify({'ok': False, 'error': 'No autorizado'}), 403
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No se envió archivo'}), 400
    f = request.files['file']
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.png', '.svg', '.jpg', '.jpeg', '.webp'):
        return jsonify({'ok': False, 'error': 'Formato no válido (png, svg, jpg, webp)'}), 400
    os.makedirs(LOGOS_DIR, exist_ok=True)
    import time as _t
    fname = 'logo_' + str(int(_t.time())) + ext
    f.save(os.path.join(LOGOS_DIR, fname))
    url = '/api/branding/logo/' + fname
    with lock:
        _state.setdefault('branding', {})['logo_url'] = url
        save_state()
    return jsonify({'ok': True, 'url': url})

@app.route('/api/branding/logo/<path:filename>')
def serve_logo(filename):
    from flask import send_from_directory
    return send_from_directory(LOGOS_DIR, filename)

@app.route('/api/pagos', methods=['GET'])
def get_pagos():
    """Estado del pago — lo consumen el manager (apartado Pagos) y el hub (público)."""
    with lock:
        return jsonify(_pago_estado())

@app.route('/api/suca/pagos', methods=['POST'])
def set_pagos():
    if not check_suca():
        return jsonify({'ok': False, 'error': 'No autorizado'}), 403
    body = request.get_json() or {}
    with lock:
        p = _state.setdefault('pagos', {})
        if 'activo' in body:     p['activo'] = bool(body['activo'])
        if 'vence' in body:      p['vence'] = str(body['vence'] or '')
        if 'nota' in body:       p['nota'] = str(body['nota'] or '')
        if 'override' in body and str(body['override']) in ('auto', 'off', '1', '2'):
            p['override'] = str(body['override'])
        if 'monto' in body:
            try: p['monto'] = float(body['monto'])
            except (ValueError, TypeError): pass
        if 'dias_aviso' in body:
            try: p['dias_aviso'] = max(0, int(body['dias_aviso']))
            except (ValueError, TypeError): pass
        save_state()
        estado = _pago_estado()
    return jsonify({'ok': True, 'pagos': estado})


# ── Estaciones dinámicas: API ────────────────────────────────────────────────

@app.route('/api/estaciones', methods=['GET'])
def api_estaciones():
    """Estaciones para el hub (sin PINs). ?all=1 incluye las desactivadas."""
    incluir = request.args.get('all') == '1'
    out = [{'id': e['id'], 'tipo': e.get('tipo', 'caja'), 'nombre': e.get('nombre', ''),
            'icono': e.get('icono') or _icono_default(e.get('tipo')), 'activa': e.get('activa', True)}
           for e in _estaciones() if incluir or e.get('activa', True)]
    return jsonify(out)

@app.route('/api/suca/estaciones', methods=['GET'])
def suca_estaciones_get():
    if not check_suca():
        return jsonify({'ok': False, 'error': 'No autorizado'}), 403
    return jsonify(_estaciones())

@app.route('/api/suca/estaciones', methods=['POST'])
def suca_estaciones_add():
    if not check_suca():
        return jsonify({'ok': False, 'error': 'No autorizado'}), 403
    body = request.get_json() or {}
    tipo = str(body.get('tipo', 'caja'))
    if tipo not in ('caja', 'tarjetas'):
        return jsonify({'ok': False, 'error': 'Tipo inválido'}), 400
    nombre = str(body.get('nombre', '')).strip() or ('Caja' if tipo == 'caja' else 'Tarjetas')
    pin = str(body.get('pin', '1234')).strip()
    if not (pin.isdigit() and len(pin) == 4):
        return jsonify({'ok': False, 'error': 'El PIN debe tener 4 dígitos'}), 400
    icono = str(body.get('icono', '')).strip() or _icono_default(tipo)
    with lock:
        cfg = load_config()
        _ensure_estaciones(cfg)
        nuevo_id = int(cfg.get('estacion_id_counter', len(cfg['estaciones']))) + 1
        e = {'id': nuevo_id, 'tipo': tipo, 'nombre': nombre, 'pin': pin, 'activa': True, 'icono': icono}
        cfg['estaciones'].append(e)
        cfg['estacion_id_counter'] = nuevo_id
        save_config(cfg)
    return jsonify({'ok': True, 'estacion': e})

@app.route('/api/suca/estaciones/<int:eid>', methods=['PUT'])
def suca_estaciones_update(eid):
    if not check_suca():
        return jsonify({'ok': False, 'error': 'No autorizado'}), 403
    body = request.get_json() or {}
    with lock:
        cfg = load_config()
        _ensure_estaciones(cfg)
        e = next((x for x in cfg['estaciones'] if int(x['id']) == eid), None)
        if not e:
            return jsonify({'ok': False, 'error': 'Estación no encontrada'}), 404
        if 'nombre' in body:
            e['nombre'] = str(body['nombre']).strip() or e['nombre']
        if 'activa' in body:
            e['activa'] = bool(body['activa'])
        if 'icono' in body:
            e['icono'] = str(body['icono']).strip() or e.get('icono') or _icono_default(e.get('tipo'))
        if 'pin' in body:
            pin = str(body['pin']).strip()
            if not (pin.isdigit() and len(pin) == 4):
                return jsonify({'ok': False, 'error': 'El PIN debe tener 4 dígitos'}), 400
            e['pin'] = pin
        save_config(cfg)
    return jsonify({'ok': True, 'estacion': e})

@app.route('/api/suca/estaciones/<int:eid>', methods=['DELETE'])
def suca_estaciones_delete(eid):
    if not check_suca():
        return jsonify({'ok': False, 'error': 'No autorizado'}), 403
    with lock:
        cfg = load_config()
        _ensure_estaciones(cfg)
        antes = len(cfg['estaciones'])
        cfg['estaciones'] = [x for x in cfg['estaciones'] if int(x['id']) != eid]
        if len(cfg['estaciones']) == antes:
            return jsonify({'ok': False, 'error': 'Estación no encontrada'}), 404
        save_config(cfg)
    return jsonify({'ok': True})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
